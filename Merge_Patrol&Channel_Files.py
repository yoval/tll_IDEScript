# -*- coding: utf-8 -*-
"""
Created on Thu May 16 15:36:35 2024

@author: Administrator
巡店表与同比表合并
"""
import os
import pandas as pd
import numpy as np
import configparser
import glob
from datetime import datetime,timedelta
import math
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook

config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
outputs_folder = config.get('WebScript', 'outputs_folder')

now = datetime.now().strftime('%Y%m%d_%H%M')

today = datetime.now()
today_ = today.strftime("%Y%m%d")
first_day_of_month = today.strftime("%m01")
yesterday = today - timedelta(days=1)
formatted_yesterday = yesterday.strftime("%m%d")

file_pattern = f'各渠道同比销售情况_{first_day_of_month}_{formatted_yesterday}*.xlsx'
xiaoshou_file = glob.glob(os.path.join(outputs_folder, file_pattern))[0]
file_pattern = f'市场部巡店表整合_{today_}*.xlsx'
xundian_file = glob.glob(os.path.join(outputs_folder, file_pattern))[0]
print('\n',xiaoshou_file.split('\\')[-1],'\n',xundian_file.split('\\')[-1])

output_file_name = f'{outputs_folder}\\市场部巡店与门店同期收银整合表_{now}.xlsx'

xiaoshou_df = pd.read_excel(xiaoshou_file, sheet_name='底表', usecols=['南北战区', '门店编码', '流水同比', '是否存量'])

xundian_df = pd.read_excel(xundian_file, sheet_name = '门店被巡详表')
merge_df = pd.merge(xundian_df, xiaoshou_df, how='left', left_on='门店编码', right_on = '门店编码')
merge_df['下降超过50%'] = np.where(merge_df['流水同比'] < -0.5, 1, 0)
# 存量门店
cunliang_df = merge_df[(merge_df['是否存量'] == '是') & (merge_df['运营状态'] == '营业中')]
cunliang_df.head(1)
# 存量、下降超过50%、未被巡查 
cunliang_df_weixuncha = cunliang_df[(cunliang_df['是否被巡查'] == 0) & (cunliang_df['下降超过50%'] == 1)]
# 透视存量门店
cunliang_pivot_df = cunliang_df.pivot_table(index =['大区经理', '省区经理', '区域经理'],
                                values = ['门店数量', '是否被巡查', '下降超过50%'],
                                aggfunc = 'sum')
weixuncha_pivot_df = cunliang_df_weixuncha.pivot_table(index=['大区经理', '省区经理', '区域经理'],
                                values=['门店数量'],
                                aggfunc='sum')
cunliang_pivot_df['未被巡查数量'] = weixuncha_pivot_df['门店数量']
cunliang_pivot_df = cunliang_pivot_df.rename(columns={'门店数量': '存量门店数',
                                                     '下降超过50%':'下降超过50%门店数',
                                                      '未被巡查数量':'下降超过50%未被巡查门店数'
                                                     })
cunliang_pivot_df = cunliang_pivot_df.loc[:, ['存量门店数', '下降超过50%门店数', '下降超过50%未被巡查门店数']]
filtered_df = merge_df[merge_df['运营状态'] == '营业中']
pivot_df = filtered_df.pivot_table(index=['大区经理', '省区经理', '区域经理'],
                                values=['门店数量', '是否被巡查', '是否交叉巡店', '是否自查'],
                                aggfunc='sum')
#增加私货
pivot_df['存量门店数'] = cunliang_pivot_df['存量门店数']
pivot_df['下降超过50%门店数'] = cunliang_pivot_df['下降超过50%门店数']
pivot_df['下降超过50%未被巡查门店数'] = cunliang_pivot_df['下降超过50%未被巡查门店数']


pivot_df = pivot_df.reset_index()
pivot_df['被巡查占比'] = pivot_df['是否被巡查']/pivot_df['门店数量']
pivot_df['被交叉巡查占比'] = pivot_df['是否交叉巡店']/pivot_df['门店数量']
pivot_df['自查占比'] = pivot_df['是否自查']/pivot_df['门店数量']
pivot_df.rename(columns={
    '是否交叉巡店': '被交叉巡店数量',
    '是否自查':'自查数量',
    '是否被巡查':'被巡查数量'
    }, inplace=True)

pivot_df = pivot_df.loc[:, ['大区经理', '省区经理', '区域经理', '门店数量','被巡查数量','被巡查占比','自查数量','自查占比','被交叉巡店数量','被交叉巡查占比','存量门店数','下降超过50%门店数','下降超过50%未被巡查门店数']]

pivot_df['距40%相差门店数'] = (pivot_df['门店数量'] * 0.4 - pivot_df['被交叉巡店数量']).apply(lambda x: math.ceil(x)).astype(int)
pivot_df['距40%相差门店数'] = pivot_df['距40%相差门店数'].clip(lower=0, upper=10000)
# 增加汇总行
result = pivot_df.pivot_table(index=["大区经理", "省区经理", "区域经理"], aggfunc="sum")
summary_by_daqu_manager = result.groupby(level="大区经理").sum().reset_index()
summary_by_daqu_manager['被巡查占比'] = summary_by_daqu_manager['被巡查数量']/summary_by_daqu_manager['门店数量']
summary_by_daqu_manager['自查占比'] = summary_by_daqu_manager['自查数量']/summary_by_daqu_manager['门店数量']
summary_by_daqu_manager['被交叉巡查占比'] = summary_by_daqu_manager['被交叉巡店数量']/summary_by_daqu_manager['门店数量']

summary_by_sheng_manager = result.groupby(level=["大区经理", "省区经理"]).sum().reset_index()
summary_by_sheng_manager['被巡查占比'] = summary_by_sheng_manager['被巡查数量']/summary_by_sheng_manager['门店数量']
summary_by_sheng_manager['自查占比'] = summary_by_sheng_manager['自查数量']/summary_by_sheng_manager['门店数量']
summary_by_sheng_manager['被交叉巡查占比'] = summary_by_sheng_manager['被交叉巡店数量']/summary_by_sheng_manager['门店数量']

pivot_df = pd.concat([pivot_df, summary_by_daqu_manager], axis=0, ignore_index=True)
pivot_df = pd.concat([pivot_df, summary_by_sheng_manager], axis=0, ignore_index=True)
pivot_df = pivot_df.sort_values(['大区经理', '省区经理', '区域经理'], ascending=[True, True, True])

# 查找“区域经理”列的空值并根据“省区经理”列是否为空进行修改
pivot_df["区域经理"] = pivot_df.apply(
    lambda row: "大区合计" if pd.isna(row["区域经理"]) and pd.isna(row["省区经理"]) else "省区合计" if pd.isna(row["区域经理"]) else row["区域经理"],
    axis=1,
)
pivot_df = pivot_df.loc[:, ['大区经理', '省区经理', '区域经理', '门店数量','被巡查数量','被巡查占比','自查数量','自查占比','被交叉巡店数量','被交叉巡查占比','距40%相差门店数','存量门店数','下降超过50%门店数','下降超过50%未被巡查门店数']]
# pivot_df.fillna(0,inplace=True)
merge_df['是否交叉巡店'] = merge_df['是否交叉巡店'].replace({1: '是', 0: '否'})
merge_df['是否自查'] = merge_df['是否自查'].replace({1: '是', 0: '否'})
merge_df['是否被巡查'] = merge_df['是否被巡查'].replace({1: '是', 0: '否'})

#复制区域经理表
quyu_df = pd.read_excel(xundian_file,sheet_name='经理巡店次数')
if not os.path.exists(output_file_name):
    # 如果文件不存在，创建一个新的Excel文件
    workbook = Workbook()
    workbook.save(output_file_name)
writer = pd.ExcelWriter(output_file_name, mode='a', engine='openpyxl', if_sheet_exists='replace')
# writer = pd.ExcelWriter(xundian_file, mode='a', engine='openpyxl', if_sheet_exists='new')
merge_df.to_excel(writer, sheet_name='门店被巡详表', index=False)
pivot_df.to_excel(writer, sheet_name='营业中门店被巡', index=False)
quyu_df.to_excel(writer, sheet_name='经理巡店次数', index=False)
# 保存更改并关闭ExcelWriter
# writer.save()
writer.close()
# 设置格式
sheng_fill_color = 'EEECE1'
daqu_fill_color = '948A54'
workbook = openpyxl.load_workbook(output_file_name) 
for worksheet in workbook.worksheets:
    for column in worksheet.columns:
        if '占比' in column[0].value or '同比' in column[0].value:
            for cell in column:
                cell.number_format = '0.00%'
    for row in worksheet.iter_rows():
        if row[2].value == "省区合计":
            for cell in row[0:18]: 
                cell.fill = PatternFill(start_color=sheng_fill_color, end_color=sheng_fill_color, fill_type="solid")
        elif row[2].value == "大区合计":
            for cell in row[0:18]: 
                cell.fill = PatternFill(start_color=daqu_fill_color, end_color=daqu_fill_color, fill_type="solid")
workbook.save(output_file_name)