# -*- coding: utf-8 -*-
"""
Created on Thu May 16 14:20:25 2024

@author: Administrator
各渠道脚本 处理结果合并
"""
import pandas as pd
import numpy as np
import glob,os,time
import configparser
from datetime import datetime,timedelta
import openpyxl
from openpyxl.styles import PatternFill


config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
outputs_folder = config.get('WebScript', 'outputs_folder')
manager_order_str = config.get('DEFAULT', 'manager_order')
manager_order = [name.strip() for name in manager_order_str.split(',')]
manager_type = pd.CategoricalDtype(categories=manager_order, ordered=True)

now = time.strftime('%Y%m%d_%H%M', time.localtime())
today = datetime.now()
first_day_of_month = today.strftime("%Y%m01")
yesterday = today - timedelta(days=1)
formatted_yesterday = yesterday.strftime("%Y%m%d")
file_pattern = f'*{first_day_of_month}_{formatted_yesterday}*.xlsx'

output_filename = f'{outputs_folder}//各渠道同比销售情况_{first_day_of_month[4:]}_{formatted_yesterday[4:]}.xlsx'
benqi_file = glob.glob(os.path.join(outputs_folder, file_pattern))[0]
last_year_file_pattern = file_pattern.replace('2024', '2023')
tongqi_file = glob.glob(os.path.join(outputs_folder, last_year_file_pattern))[0]
benqi_df = pd.read_excel(benqi_file)
tongqi_df = pd.read_excel(tongqi_file)
tongqi_df = tongqi_df.loc[:,['门店编码','时段','营业天数','流水金额','实收金额','订单数','堂食流水','堂食实收','堂食订单数','外卖流水','外卖实收','外卖订单数','小程序流水','小程序实收','小程序订单数']]
df_merge= pd.merge(benqi_df, tongqi_df, how='left', left_on='门店编码',right_on='门店编码',suffixes=('_本期', '_同期')).fillna(0)
df_merge['是否存量'] = df_merge.apply(lambda row: '否' if row['流水金额_本期'] * row['流水金额_同期'] == 0 else '是', axis=1)
df_merge['流水同比'] = (df_merge['流水金额_本期'] - df_merge['流水金额_同期'])/df_merge['流水金额_同期']
df_merge['堂食流水同比'] = (df_merge['堂食流水_本期'] - df_merge['堂食流水_同期'])/df_merge['堂食流水_同期']
df_merge['外卖流水同比'] = (df_merge['外卖流水_本期'] - df_merge['外卖流水_同期'])/df_merge['外卖流水_同期']
df_merge['小程序流水同比'] = (df_merge['小程序流水_本期'] - df_merge['小程序流水_同期'])/df_merge['小程序流水_同期']
df_merge['订单数同比'] = (df_merge['订单数_本期'] - df_merge['订单数_同期'])/df_merge['订单数_同期']
df_merge['堂食订单数同比'] = (df_merge['堂食订单数_本期'] - df_merge['堂食订单数_同期'])/df_merge['堂食订单数_同期']
df_merge['外卖订单数同比'] = (df_merge['外卖订单数_本期'] - df_merge['外卖订单数_同期'])/df_merge['外卖订单数_同期']
df_merge['小程序订单数同比'] = (df_merge['小程序订单数_本期'] - df_merge['小程序订单数_同期'])/df_merge['小程序订单数_同期']
df_merge['实收金额同比'] = (df_merge['实收金额_本期'] - df_merge['实收金额_同期'])/df_merge['实收金额_同期']
df_merge['堂食实收同比'] = (df_merge['堂食实收_本期'] - df_merge['堂食实收_同期'])/df_merge['堂食实收_同期']
df_merge['外卖实收同比'] = (df_merge['外卖实收_本期'] - df_merge['外卖实收_同期'])/df_merge['外卖实收_同期']
df_merge['小程序实收同比'] = (df_merge['小程序实收_本期'] - df_merge['小程序实收_同期'])/df_merge['小程序实收_同期']
df_merge['订单价_本期'] = df_merge['流水金额_本期'] / df_merge['订单数_本期']
df_merge['订单价_同期'] = df_merge['流水金额_同期'] / df_merge['订单数_同期']
df_merge['订单价增减'] = df_merge['订单价_本期'] - df_merge['订单价_同期'] 
df_merge = df_merge.replace([np.inf, -np.inf], 0)

df_merge = df_merge.loc[:,['门店编码','门店名称','大区经理','省区经理','区域经理','南北战区','运营状态','省','市','区','U8C客商编码',
                            '时段_本期','时段_同期',
                            '营业天数_本期','营业天数_同期',
                            '流水金额_本期','流水金额_同期','流水同比',
                            '实收金额_本期','实收金额_同期','实收金额同比',
                            '订单数_本期','订单数_同期','订单数同比',
                            '订单价_本期','订单价_同期','订单价增减',
                            '堂食流水_本期','堂食流水_同期','堂食流水同比',
                            '堂食实收_本期','堂食实收_同期','堂食实收同比',
                            '堂食订单数_本期','堂食订单数_同期','堂食订单数同比',
                            '外卖流水_本期','外卖流水_同期','外卖流水同比',
                            '外卖实收_本期','外卖实收_同期','外卖实收同比',
                            '外卖订单数_本期','外卖订单数_同期','外卖订单数同比',
                            '小程序流水_本期','小程序流水_同期','小程序流水同比',
                            '小程序实收_本期','小程序实收_同期','小程序实收同比',
                            '小程序订单数_本期','小程序订单数_同期','小程序订单数同比',
                            '是否存量']
]


# 存量
df_cunliang = df_merge[(df_merge['运营状态'] == '营业中') & (df_merge['是否存量'] == '是')]
df_cunliang['上升数']  = (df_cunliang["流水金额_本期"] > df_cunliang["流水金额_同期"]).astype(int)
df_cunliang['下降数'] = (df_cunliang["流水金额_本期"] < df_cunliang["流水金额_同期"]).astype(int)

# 计算门店编号的计数
store_count = pd.pivot_table(
    df_cunliang,
    index=["大区经理", "省区经理", "区域经理"],
    values=["门店编码"],
    aggfunc="count"
)

# 流水
pivot_liushui = df_cunliang.pivot_table(index=['大区经理', '省区经理', '区域经理'],
                                       values=['上升数','下降数','流水金额_本期', '流水金额_同期', '堂食流水_本期', '堂食流水_同期','外卖流水_本期','外卖流水_同期','小程序流水_本期','小程序流水_同期'],
                                       aggfunc='sum')

pivot_liushui['门店数量'] = store_count['门店编码']
summary_by_daqu_manager = pivot_liushui.groupby(level="大区经理").sum().reset_index()
summary_by_sheng_manager = pivot_liushui.groupby(level=["大区经理", "省区经理"]).sum().reset_index()

pivot_liushui = pivot_liushui.reset_index()
pivot_liushui = pd.concat([pivot_liushui, summary_by_daqu_manager], axis=0, ignore_index=True)
pivot_liushui = pd.concat([pivot_liushui, summary_by_sheng_manager], axis=0, ignore_index=True)

pivot_liushui['流水同比'] = (pivot_liushui['流水金额_本期'] - pivot_liushui['流水金额_同期'])/pivot_liushui['流水金额_同期']
pivot_liushui['堂食流水同比'] = (pivot_liushui['堂食流水_本期'] - pivot_liushui['堂食流水_同期'])/pivot_liushui['堂食流水_同期']
pivot_liushui['外卖流水同比'] = (pivot_liushui['外卖流水_本期'] - pivot_liushui['外卖流水_同期'])/pivot_liushui['外卖流水_同期']
pivot_liushui['小程序流水同比'] = (pivot_liushui['小程序流水_本期'] - pivot_liushui['小程序流水_同期'])/pivot_liushui['小程序流水_同期']
pivot_liushui = pivot_liushui.loc[:,['大区经理','省区经理','区域经理','门店数量','上升数','下降数','流水金额_本期', '流水金额_同期','流水同比', '堂食流水_本期', '堂食流水_同期','堂食流水同比','外卖流水_本期','外卖流水_同期','外卖流水同比','小程序流水_本期','小程序流水_同期','小程序流水同比']]
pivot_liushui['大区经理'] = pivot_liushui['大区经理'].astype(manager_type)
pivot_liushui = pivot_liushui.sort_values(['大区经理', '省区经理', '区域经理'], ascending= True)
# 查找“区域经理”列的空值并根据“省区经理”列是否为空进行修改
pivot_liushui["区域经理"] = pivot_liushui.apply(
    lambda row: "大区合计" if pd.isna(row["区域经理"]) and pd.isna(row["省区经理"]) else "省区合计" if pd.isna(row["区域经理"]) else row["区域经理"],
    axis=1,
)


# 单量
pivot_dingdan = df_cunliang.pivot_table(index=['大区经理', '省区经理', '区域经理'],
                                       values=['订单数_本期', '订单数_同期', '堂食订单数_本期', '堂食订单数_同期','外卖订单数_本期','外卖订单数_同期','小程序订单数_本期','小程序订单数_同期'],
                                       aggfunc='sum')

pivot_dingdan['门店数量'] = store_count['门店编码']

summary_by_daqu_manager = pivot_dingdan.groupby(level="大区经理").sum().reset_index()
summary_by_sheng_manager = pivot_dingdan.groupby(level=["大区经理", "省区经理"]).sum().reset_index()

pivot_dingdan = pivot_dingdan.reset_index()
pivot_dingdan = pd.concat([pivot_dingdan, summary_by_daqu_manager], axis=0, ignore_index=True)
pivot_dingdan = pd.concat([pivot_dingdan, summary_by_sheng_manager], axis=0, ignore_index=True)

pivot_dingdan['订单数同比'] = (pivot_dingdan['订单数_本期'] - pivot_dingdan['订单数_同期'])/pivot_dingdan['订单数_同期']
pivot_dingdan['堂食订单数同比'] = (pivot_dingdan['堂食订单数_本期'] - pivot_dingdan['堂食订单数_同期'])/pivot_dingdan['堂食订单数_同期']
pivot_dingdan['外卖订单数同比'] = (pivot_dingdan['外卖订单数_本期'] - pivot_dingdan['外卖订单数_同期'])/pivot_dingdan['外卖订单数_同期']
pivot_dingdan['小程序订单数同比'] = (pivot_dingdan['小程序订单数_本期'] - pivot_dingdan['小程序订单数_同期'])/pivot_dingdan['小程序订单数_同期']
pivot_dingdan = pivot_dingdan.loc[:,['大区经理','省区经理','区域经理','门店数量','订单数_本期', '订单数_同期','订单数同比', '堂食订单数_本期', '堂食订单数_同期','堂食订单数同比','外卖订单数_本期','外卖订单数_同期','外卖订单数同比','小程序订单数_本期','小程序订单数_同期','小程序订单数同比']]
pivot_dingdan['大区经理'] = pivot_dingdan['大区经理'].astype(manager_type)
pivot_dingdan = pivot_dingdan.sort_values(['大区经理', '省区经理', '区域经理'], ascending= True)
# 查找“区域经理”列的空值并根据“省区经理”列是否为空进行修改
pivot_dingdan["区域经理"] = pivot_dingdan.apply(
    lambda row: "大区合计" if pd.isna(row["区域经理"]) and pd.isna(row["省区经理"]) else "省区合计" if pd.isna(row["区域经理"]) else row["区域经理"],
    axis=1,
)

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    df_merge.to_excel(writer, sheet_name='底表', index=False)
    pivot_liushui.to_excel(writer, sheet_name='营业中存量门店流水', index=False)
    pivot_dingdan.to_excel(writer, sheet_name='营业中存量门店单量', index=False)



#格式设置
sheng_fill_color = 'EEECE1'
daqu_fill_color = '948A54'

# 加载Excel文件
workbook = openpyxl.load_workbook(output_filename) 
# 遍历所有工作表
for worksheet in workbook.worksheets:
    for column in worksheet.columns:
        if ("环比" in column[0].value or "同比" in column[0].value or "同比" in column[0].value) and ("期" not in column[0].value):
            for cell in column:
                cell.number_format = '0.00%'
        elif "流水" in column[0].value:
            for cell in column:
                cell.number_format = '0"."0,"万"'
    for row in worksheet.iter_rows():
        if row[2].value == "省区合计":
            for cell in row[0:18]: 
                cell.fill = PatternFill(start_color=sheng_fill_color, end_color=sheng_fill_color, fill_type="solid")
        elif row[2].value == "大区合计":
            for cell in row[0:18]: 
                cell.fill = PatternFill(start_color=daqu_fill_color, end_color=daqu_fill_color, fill_type="solid")
workbook.save(output_filename)